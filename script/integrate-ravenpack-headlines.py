#!/usr/bin/env python3

"""
Replace the sha256 hashes with the correct RavenPact headlines.

Usage:
    Copy `news_headlines_and_sentiment_cogn2022.xlsx` and a file with the
    headlines retrieved from RavenPact (`headlines.xlsx`) into the
    current directory.

    The `integrate-ravenpack-headlines.py` script will then replace hashes
    for which a headline could be found with the corresponding headline.
"""

from openpyxl import load_workbook
from hashlib import sha256
from pathlib import Path
from sys import exit

dataset_fname = 'news_headlines_and_sentiment_cogn2022.xlsx'
headline_fname = 'headlines.xlsx'

#
# Verify that all necessary files are available
#
if not Path(dataset_fname).is_file():
    print(f'Cannot find the datset file `{dataset_fname}` in the current '
          f'working directory.')
    exit(-1)

if not Path(headline_fname).is_file():
    print(f'Cannot find the headlines file `{headline_fname}` in the current '
          f'working directory.')
    exit(-1)

#
# create the mapping between headlines and hashes
#
wb = load_workbook(filename=headline_fname)
ws = wb['Sheet1']

hash_to_headline = {}
i = 2
while value := ws[f'C{i}'].value:
    text = value.strip()
    hash_to_headline[sha256(text.encode('utf8')).hexdigest()] = text
    i += 1

#
# replace known hashes with the original title
#
wb = load_workbook(filename=dataset_fname)
ws = wb['Sheet1']

i = 2
while value := ws[f'B{i}'].value:
    text = value.strip()
    if text in hash_to_headline:
        ws[f'B{i}'] = hash_to_headline[text]
    i += 1

wb.save(dataset_fname.split('.')[0] + '_headlines.xlsx')
